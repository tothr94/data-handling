import openpyxl
from openpyxl import Workbook

from data.basic import generator
from data.basic.model_dataclasses import Person, Car, Airport


def write_people(people: list[Person],
                 workbook: openpyxl.Workbook,
                 sheet_name: str = "people",
                 heading: bool = True) -> None:
    sheet = workbook.create_sheet(sheet_name if sheet_name is not None else "people")

    if heading:
        field_names = ["id", "name", "age", "male"]
        for col in range(len(field_names)):
            sheet.cell(row=1, column=col + 1, value=field_names[col])

    offset = 2 if heading else 1
    for row in range(len(people)):
        sheet.cell(row=row + offset, column=1, value=people[row].id)
        sheet.cell(row=row + offset, column=2, value=people[row].name)
        sheet.cell(row=row + offset, column=3, value=people[row].age)
        sheet.cell(row=row + offset, column=4, value=people[row].male)


def read_people(workbook: openpyxl.Workbook,
                sheet_name: str = "people") -> list[Person]:
    sheet = workbook[sheet_name]

    people = []
    row = 1
    # for row in range(1, 11):
    while True:
        if sheet.cell(row=row, column=1).value is None:
            break
        person = Person(
            sheet.cell(row=row, column=1).value,
            sheet.cell(row=row, column=2).value,
            int(sheet.cell(row=row, column=3).value),
            bool(sheet.cell(row=row, column=4).value)
        )
        row += 1
        people.append(person)
    return people


def write_cars(cars: list[Car],
               workbook: openpyxl.Workbook,
               sheet_name: str = "cars",
               heading: bool = True) -> None:
    sheet = workbook.create_sheet(sheet_name if sheet_name is not None else "cars")

    if heading:
        field_names = ["plate", "type", "year", "automatic"]
        for col in range(len(field_names)):
            sheet.cell(row=1, column=col + 1, value=field_names[col])

    offset = 2 if heading else 1
    for row in range(len(cars)):
        sheet.cell(row=row + offset, column=1, value=cars[row].plate)
        sheet.cell(row=row + offset, column=2, value=cars[row].type)
        sheet.cell(row=row + offset, column=3, value=cars[row].year)
        sheet.cell(row=row + offset, column=4, value=cars[row].automatic)


def read_cars(workbook: openpyxl.Workbook,
              sheet_name: str = "cars",
              heading: bool = True) -> list[Car]:
    sheet = workbook[sheet_name if sheet_name is not None else "cars"]
    cars = []

    row = 2 if heading else 1
    while True:
        cell = sheet.cell(row=row, column=1)
        if cell.value is None:
            break

        cars.append(Car(
            sheet.cell(row=row, column=1).value,
            sheet.cell(row=row, column=2).value,
            int(sheet.cell(row=row, column=3).value),
            bool(sheet.cell(row=row, column=4).value)
        ))
        row += 1

    return cars


def write_airports(airports: list[Airport],
                   workbook: openpyxl.Workbook,
                   sheet_name: str = "airports",
                   heading: bool = True) -> None:
    sheet = workbook.create_sheet(sheet_name if sheet_name is not None else "airports")

    if heading:
        field_names = ["code", "name", "city", "state", "country"]
        for col in range(len(field_names)):
            sheet.cell(row=1, column=col + 1, value=field_names[col])

    offset = 2 if heading else 1
    for row in range(len(airports)):
        sheet.cell(row=row + offset, column=1, value=airports[row].code)
        sheet.cell(row=row + offset, column=2, value=airports[row].name)
        sheet.cell(row=row + offset, column=3, value=airports[row].city)
        sheet.cell(row=row + offset, column=4, value=airports[row].state)
        sheet.cell(row=row + offset, column=5, value=airports[row].country)


def read_airports(workbook: openpyxl.Workbook,
                  sheet_name: str = "airports",
                  heading: bool = True) -> list[Airport]:
    sheet = workbook[sheet_name if sheet_name is not None else "airports"]
    airports = []

    row = 2 if heading else 1
    while True:
        cell = sheet.cell(row=row, column=1)
        if cell.value is None:
            break

        airports.append(Airport(
            sheet.cell(row=row, column=1).value,
            sheet.cell(row=row, column=2).value,
            sheet.cell(row=row, column=3).value,
            sheet.cell(row=row, column=4).value,
            sheet.cell(row=row, column=4).value
        ))
        row += 1

    return airports


if __name__ == "__main__":
    wb = Workbook()
    # wb = openpyxl.load_workbook("C:/hallgato/data.xlsx")
    write_people(generator.generate_people(20), wb)
    for person in read_people(wb):
        print(person)
    wb.save("C:/hallgato/data.xlsx")
